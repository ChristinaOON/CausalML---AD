### Part One - Data import, data processing, data filling, and filling effect evaluation ###

## Fill in missing data values ##

from sklearn.experimental import enable_iterative_imputer
from sklearn.impute import IterativeImputer
from sklearn.neighbors import KNeighborsRegressor
from sklearn.ensemble import RandomForestRegressor
from lightgbm import LGBMRegressor
import pandas as pd

# Data import
data_na = pd.read_csv("jz_data.csv")

# Define three models to use to populate missing values
imputers = [KNeighborsRegressor, RandomForestRegressor, LGBMRegressor]

# Create an empty list to store the populated data
imputed_data = []

for imputer in imputers:
    # Initialize the iterative filling model, specifying the filling model, the maximum number of iterations, and the random state
    imp = IterativeImputer(estimator=imputer(), max_iter=100, random_state=1)
    # Add the populated data to the list
    filled_data = pd.DataFrame(imp.fit_transform(data_na), columns=data_na.columns)
    imputed_data.append(filled_data)

# Save the data for each filling method as a CSV file
for i, data in enumerate(imputed_data):
    imputer_name = imputers[i].__name__
    data.to_csv(f'imputed_data_{imputer_name}.csv', index=False)

## Evaluation of the effect of filling ##

from autogluon.tabular import TabularDataset, TabularPredictor
import numpy as np
import pandas as pd
import shap
shap.initjs()
import time
import warnings
warnings.filterwarnings('ignore')
from matplotlib import pyplot as plt
from sklearn.model_selection import train_test_split
import seaborn as sns
import sklearn
import sklearn.model_selection
from sklearn.metrics import r2_score
from PIL import Image
import os

data_knn=pd.read_csv('imputed_data_KNeighborsRegressor.csv')
data_lgb=pd.read_csv('imputed_data_LGBMRegressor.csv')
data_rf=pd.read_csv('imputed_data_RandomForestRegressor.csv')

# Define independent and dependent variables
features = ['PWFV','FW','ICO','AO','Manure','Percolate','TFV','TSf','VSf','HFV','HRT','Addition','Temperature','OLR','VFA','pH','TSo','VSo','TAN']  # 替换为您的自变量列名
labels = ['CH4', 'H2S', 'Biogas_Production']

# Define the division ratio between the training set and the test set
test_size = 0.3

# Loop through each dataset and dependent variable
for dataset, dataset_name in zip([data_knn, data_lgb, data_rf], ['knn', 'lgb', 'rf']):
    for label in labels:
        # Extract the independent and dependent variables
        X = dataset[features]
        y = dataset[label]
        
        # Divide the training set and the test set
        X_train, X_test, y_train, y_test = sklearn.model_selection.train_test_split(X, y, test_size=test_size, random_state=0)
        
        # Build the DataFrame of the training set
        train_data = X_train.copy()
        train_data[label] = y_train
        
        test_data = X_test.copy()
        test_data[label] = y_test
        # Create an empty DataFrame to store the test results
        results = pd.DataFrame(columns=['best_model', 'train_rmse', 'test_rmse','train_r2','test_r2'])

        # Use AutoGluon for model training
        predictor = TabularPredictor(label=label, problem_type='regression')
        predictor.fit(train_data, time_limit=60,presets='best_quality',auto_stack=True)

        # Get the optimal model name
        best_model = predictor.leaderboard(silent=True)['model'].iloc[0]

        # Root mean square error (RMSE) and R2 scores were calculated on the training and test sets
        train_rmse = predictor.evaluate(train_data, silent=True)['root_mean_squared_error']
        test_rmse = predictor.evaluate(test_data, silent=True)['root_mean_squared_error']
        train_r2 = r2_score(train_data[label], predictor.predict(train_data))
        test_r2 = r2_score(test_data[label], predictor.predict(test_data))

        # Add the result to the DataFrame
        results = results.append({'best_model': best_model, 'train_rmse': train_rmse, 'test_rmse': test_rmse, 'train_r2': train_r2, 'test_r2': test_r2}, ignore_index=True)

        # Print the results
        # Print the dataset name, the dependent variable name, the optimal model name, and the evaluation index value
        print(f'数据集：{dataset_name}，因变量：{label}，最优模型：{results["best_model"].iloc[0]}')
        print(f'在训练集上的RMSE:{results["train_rmse"].iloc[0]}') 
        print(f'在训练集上的R2:{results["train_r2"].iloc[0]}') 
        print(f'在测试集上的RMSE:{results["test_rmse"].iloc[0]}') 
        print(f'在测试集上的R2:{results["test_r2"].iloc[0]}')

### The three evaluation effects were compared, and the best evaluation model was KNN, and the final output data file format was manually changed to Excel, that is, the "jzdataknn.xlsx" data will be used for model training and machine learning research in the future ###
### Part Two - Predicts intermediate parameters and gas production performance parameters in two stages, and selects the best prediction model for four variables: TAN, VFA, methane content, and biogas yield from seven machine learning algorithm models ###
import pandas as pd
import numpy as np
# Cross-validation and hyperparameter optimization
import optuna 
from sklearn.model_selection import train_test_split  # Splitting dataset
from sklearn.model_selection import cross_validate, KFold 
# Functions for model performance evaluation metrics
from sklearn.metrics import mean_squared_error, r2_score  # MSE, MAE, R2
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, Lasso
from sklearn.svm import SVR
from sklearn.neighbors import KNeighborsRegressor
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor, StackingRegressor
from xgboost import XGBRegressor
import lightgbm as lgb
from catboost import CatBoostRegressor
from matplotlib import pyplot as plt
plt.rcParams['font.sans-serif'] =['Times New Roman']


## --------- TAN ---------- ##
## Data segmentation and normalization ##
def preprocess_data_tan (file_path):

    file_path = 'jz_data_knn.xlsx'
    data = pd.read_excel(file_path)

    X = data[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'TSf', 'VSf', 'HFV', 'HRT', 'Addition', 'Temperature', 'OLR' , 'TSo', 'VSo', 'pH']]
    y = data[['TAN']]    

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=0)

    scaler_X = StandardScaler()
    X_train = scaler_X.fit_transform(X_train)
    X_test = scaler_X.transform(X_test)

    scaler_y = StandardScaler()
    y_train = scaler_y.fit_transform(y_train.values.reshape(-1, 1)).flatten()
    y_test = scaler_y.transform(y_test.values.reshape(-1, 1)).flatten()

    return X_train, X_test, y_train, y_test, scaler_X, scaler_y

def RMSE(y, y_pred):
    RMSE = np.sqrt(mean_squared_error(y, y_pred))
    return RMSE

def MAPE(y, y_pred):
    e = np.abs((y_pred - y) / y)
    MAPE = np.sum(e) / len(e)
    return MAPE

## Predictive model definition and parameter tuning ##
# Define models
models = {
    'SVR': SVR(),
    'KNN': KNeighborsRegressor(),
    'RandomForest': RandomForestRegressor(),
    'GradientBoosting': GradientBoostingRegressor(),
    'XGBoost': XGBRegressor(),
    'LightGBM': lgb.LGBMRegressor(),
    'CatBoost': CatBoostRegressor()
}

# Hyperparameter Tuning and Model Optimization with Bayesian Optimization
# Step 1: Define the Objective Function and Parameter Space
def optuna_objective(trial, model_type, x_train, y_train):
    if model_type == 'SVR':
        C = trial.suggest_loguniform("C", 1e-3, 1e3)
        epsilon = trial.suggest_float("epsilon", 0.01, 0.5)
        kernel = trial.suggest_categorical("kernel", ['linear', 'poly', 'rbf', 'sigmoid'])
        degree = trial.suggest_int("degree", 1, 5)
        gamma = trial.suggest_categorical("gamma", ['scale', 'auto'])
        model = SVR(C=C, epsilon=epsilon, kernel=kernel, degree=degree, gamma=gamma)

    elif model_type == 'KNN':
        n_neighbors = trial.suggest_int("n_neighbors", 1, 30)
        weights = trial.suggest_categorical("weights", ['uniform', 'distance'])
        p = trial.suggest_int("p", 1, 2)
        model = KNeighborsRegressor(n_neighbors=n_neighbors, weights=weights, p=p)
    
    elif model_type == 'RandomForest':
        n_estimators = trial.suggest_int("n_estimators", 10, 1000)
        max_depth = trial.suggest_int("max_depth", 1, 20)
        min_samples_split = trial.suggest_int("min_samples_split", 2, 20)
        min_samples_leaf = trial.suggest_int("min_samples_leaf", 1, 20)
        model = RandomForestRegressor(n_estimators=n_estimators, max_depth=max_depth, 
                                      min_samples_split=min_samples_split, min_samples_leaf=min_samples_leaf)
    
    elif model_type == 'GradientBoosting':
        n_estimators = trial.suggest_int("n_estimators", 10, 1000)
        learning_rate = trial.suggest_float("learning_rate", 0.001, 1.0)
        max_depth = trial.suggest_int("max_depth", 1, 20)
        min_samples_split = trial.suggest_int("min_samples_split", 2, 20)
        min_samples_leaf = trial.suggest_int("min_samples_leaf", 1, 20)
        model = GradientBoostingRegressor(n_estimators=n_estimators, learning_rate=learning_rate, 
                                           max_depth=max_depth, min_samples_split=min_samples_split, 
                                           min_samples_leaf=min_samples_leaf)
    
    elif model_type == 'XGBoost':
        max_depth = trial.suggest_int("max_depth", 1, 20)
        learning_rate = trial.suggest_float("learning_rate", 0.01, 1.0)
        n_estimators = trial.suggest_int("n_estimators", 10, 1000)
        subsample = trial.suggest_float("subsample", 0.3, 1.0)
        colsample_bytree = trial.suggest_float("colsample_bytree", 0.3, 1.0)
        reg_alpha = trial.suggest_float("reg_alpha", 0.0, 1.0)
        reg_lambda = trial.suggest_float("reg_lambda", 0.0, 1.0)
        model = XGBRegressor(max_depth=max_depth, learning_rate=learning_rate, n_estimators=n_estimators, 
                             subsample=subsample, colsample_bytree=colsample_bytree, 
                             reg_alpha=reg_alpha, reg_lambda=reg_lambda, objective='reg:squarederror')
    
    elif model_type == 'LightGBM':
        num_leaves = trial.suggest_int("num_leaves", 2, 100)
        learning_rate = trial.suggest_float("learning_rate", 0.01, 1.0)
        n_estimators = trial.suggest_int("n_estimators", 10, 1000)
        min_child_samples = trial.suggest_int("min_child_samples", 1, 20)
        subsample = trial.suggest_float("subsample", 0.3, 1.0)
        colsample_bytree = trial.suggest_float("colsample_bytree", 0.3, 1.0)
        reg_alpha = trial.suggest_float("reg_alpha", 0.0, 1.0)
        reg_lambda = trial.suggest_float("reg_lambda", 0.0, 1.0)
        model = lgb.LGBMRegressor(num_leaves=num_leaves, learning_rate=learning_rate, n_estimators=n_estimators, 
                                  min_child_samples=min_child_samples, subsample=subsample, 
                                  colsample_bytree=colsample_bytree, reg_alpha=reg_alpha, reg_lambda=reg_lambda)

    elif model_type == 'CatBoost':
        iterations = trial.suggest_int("iterations", 10, 500)
        learning_rate = trial.suggest_float("learning_rate", 1e-5, 0.1, log=True)
        depth = trial.suggest_int("depth", 2, 10)
        l2_leaf_reg = trial.suggest_float("l2_leaf_reg", 1e-2, 0.1, log=True)
        border_count = trial.suggest_int("border_count", 1, 200)
        random_strength = trial.suggest_float("random_strength", 0.0, 1.0)
        colsample_bylevel = trial.suggest_float("colsample_bylevel", 0.01, 1.0)
        model = CatBoostRegressor(iterations=iterations, learning_rate=learning_rate, depth=depth, 
                                  l2_leaf_reg=l2_leaf_reg, border_count=border_count, 
                                  random_strength=random_strength, colsample_bylevel=colsample_bylevel, silent=True)
    
    # Cross-validation process, output negative root mean squared error (-RMSE)
    cv = KFold(n_splits=5, shuffle=True, random_state=42)
    validation_loss = cross_validate(model, x_train, y_train,
                                     scoring="neg_root_mean_squared_error",
                                     cv=cv,
                                     verbose=False,
                                     n_jobs=-1,
                                     error_score='raise')
    return np.mean(abs(validation_loss["test_score"]))

# Step 2: Define the Specific Workflow for Optimizing the Objective Function
def optimizer_optuna(X_t, y_t, n_trials, model_type):
    # Define the sampler
    algo = optuna.samplers.TPESampler(n_startup_trials=10, n_ei_candidates=24)
    # Create an Optuna optimization task
    study = optuna.create_study(sampler=algo, direction="minimize")
    # Optimize the objective function
    study.optimize(lambda trial: optuna_objective(trial, model_type, X_t, y_t), 
                   n_trials=n_trials, 
                   show_progress_bar=True 
                  )
    # Displays the best parameters and scores
    print("\n", "\n", "best params: ", study.best_trial.params,
          "\n", "\n", "best score: ", study.best_trial.value,
          "\n")
    return study.best_trial.params, study.best_trial.value

# Step 3: Train and evaluate models one by one
def train_and_evaluate_model(X_train, y_train, X_test, y_test, scaler_y):
    # Test each model and store the results
    results = {}
    model_types = {
        'KNN': KNeighborsRegressor,
        'RandomForest': RandomForestRegressor,
        'GradientBoosting': GradientBoostingRegressor,
        'SVR': SVR,
        'XGBoost': XGBRegressor,
        'LightGBM': lgb.LGBMRegressor,
        'CatBoost': CatBoostRegressor
    }
    for model_type in model_types:
        # Hyperparameter optimization
        param, best_score = optimizer_optuna(X_train, y_train, 30, model_type)
        
        # Filter the parameters and keep only the parameters that are needed for the current model
        filtered_param = {}
        for key in param:
            if key in model_types[model_type]().get_params().keys():
                filtered_param[key] = param[key]
    
        # Construct and fit the model based on the model type
        model_class = model_types[model_type]
        model = model_class(**filtered_param)
        model.fit(X_train, y_train)
        y_pred_train = model.predict(X_train)
        y_pred_test = model.predict(X_test)

        # Reverse scaling
        y_train_inverse = scaler_y.inverse_transform(y_train.reshape(-1, 1)).flatten()
        y_test_inverse = scaler_y.inverse_transform(y_test.reshape(-1, 1)).flatten()
        y_pred_train_inverse = scaler_y.inverse_transform(y_pred_train.reshape(-1, 1)).flatten()
        y_pred_test_inverse = scaler_y.inverse_transform(y_pred_test.reshape(-1, 1)).flatten()

        # Evaluate metrics
        train_r2 = r2_score(y_train_inverse, y_pred_train_inverse)
        test_r2 = r2_score(y_test_inverse, y_pred_test_inverse)
        train_rmse = np.sqrt(mean_squared_error(y_train_inverse, y_pred_train_inverse))
        test_rmse = np.sqrt(mean_squared_error(y_test_inverse, y_pred_test_inverse))

        # Store results
        results[model_type] = {
            'Model': model_type,
            'Train R2': train_r2,
            'Test R2': test_r2,
            'Train RMSE': train_rmse,
            'Test RMSE': test_rmse,
            'Best Params': filtered_param,
            'CV R2': best_score
        }

    # Choose top two models based on Test R2
    top_models = sorted(results.values(), key=lambda x: x['Test R2'], reverse=True)[:2]
    top_model_names = [model['Model'] for model in top_models]

    # Create a dictionary of the best estimators for stacking
    estimators = [(name, model_types[name](**results[name]['Best Params'])) for name in top_model_names]

    # Define final estimators for stacking
    final_estimators_combinations = {
        'Linear': LinearRegression(),
        'Ridge': Ridge(),
        'Lasso': Lasso()
    }

    # Store stacking results
    predictor_results = {}

    # Train and evaluate stacking models
    for name, final_estimator in final_estimators_combinations.items():
        predictor_model = StackingRegressor(estimators=estimators, final_estimator=final_estimator)
        predictor_model.fit(X_train, y_train)
        
        # Evaluate stacking model performance
        y_pred_train_stacking = predictor_model.predict(X_train)
        y_pred_test_stacking = predictor_model.predict(X_test)
        
        # Reverse scaling
        y_pred_train_stacking_inverse = scaler_y.inverse_transform(y_pred_train_stacking.reshape(-1, 1)).flatten()
        y_pred_test_stacking_inverse = scaler_y.inverse_transform(y_pred_test_stacking.reshape(-1, 1)).flatten()

        # Evaluate metrics
        train_r2_stacking = r2_score(y_train_inverse, y_pred_train_stacking_inverse)
        test_r2_stacking = r2_score(y_test_inverse, y_pred_test_stacking_inverse)
        train_rmse_stacking = np.sqrt(mean_squared_error(y_train_inverse, y_pred_train_stacking_inverse))
        test_rmse_stacking = np.sqrt(mean_squared_error(y_test_inverse, y_pred_test_stacking_inverse))

        # Store stacking results
        predictor_results[f'Stacking_{name}'] = {
            'Model': f'Stacking_{name}',
            'Train R2': train_r2_stacking,
            'Test R2': test_r2_stacking,
            'Train RMSE': train_rmse_stacking,
            'Test RMSE': test_rmse_stacking,
            'Best Params': ','.join(top_model_names),
            'CV R2': 'N/A'
        }

    # Convert results to DataFrame
    results_df = pd.DataFrame.from_dict(results, orient='index')
    stacking_dfs = [pd.DataFrame([value]) for key, value in predictor_results.items()]
    all_results_df = pd.concat([results_df] + stacking_dfs, ignore_index=True)
    all_results_df.sort_values(by='Test R2', ascending=False, inplace=True)

    # Update model dictionary
    trained_models = {name: model_types[name](**results[name]['Best Params']) for name in model_types}
    trained_models.update({'Stacking': predictor_model})

    # Find the best model
    best_model_name = results_df.loc[results_df['Test R2'].idxmax()]['Model']
    best_model = trained_models[best_model_name]

    return all_results_df, trained_models, best_model

## Import data and train models ##
# Data import
X_train, X_test, y_train, y_test, scaler_X, scaler_y = preprocess_data_tan (r"D:\a\jz_data_knn.xlsx")

# train and evaluate models
results_df, trained_models, best_model = train_and_evaluate_model(X_train, y_train, X_test, y_test, scaler_y)

# Convert results_df to an Excel sheet and save it
output_file = 'model_evaluation_results_TAN.xlsx'  
results_df.to_excel(output_file, index=False)  # Save to an Excel file, index=False means that the row index is not saved

print(f"Results have been saved to {output_file}")

## Intermediate parameter TAN's list of predictive machine learning effects ##
results_df

## --------- VFA ---------- ##
## Data segmentation and normalization ##
def preprocess_data_vfa (file_path):

    file_path = 'jz_data_knn.xlsx'
    data = pd.read_excel(file_path)

    X = data[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'TSf', 'VSf', 'HFV', 'HRT', 'Addition', 'Temperature', 'OLR' , 'TSo', 'VSo', 'pH']]
    y = data[['VFA']]     

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=0)

    scaler_X = StandardScaler()
    X_train = scaler_X.fit_transform(X_train)
    X_test = scaler_X.transform(X_test)

    scaler_y = StandardScaler()
    y_train = scaler_y.fit_transform(y_train.values.reshape(-1, 1)).flatten()
    y_test = scaler_y.transform(y_test.values.reshape(-1, 1)).flatten()

    return X_train, X_test, y_train, y_test, scaler_X, scaler_y

def RMSE(y, y_pred):
    RMSE = np.sqrt(mean_squared_error(y, y_pred))
    return RMSE

def MAPE(y, y_pred):
    e = np.abs((y_pred - y) / y)
    MAPE = np.sum(e) / len(e)
    return MAPE

## Import data and train models ##
# Data import
X_train, X_test, y_train, y_test, scaler_X, scaler_y = preprocess_data_vfa (r"D:\a\jz_data_knn.xlsx")

# train and evaluate models
results_df, trained_models, best_model = train_and_evaluate_model(X_train, y_train, X_test, y_test, scaler_y)

# Convert results_df to an Excel sheet and save it
output_file = 'model_evaluation_results_VFA.xlsx'  
results_df.to_excel(output_file, index=False)  # Save to an Excel file, index=False means that the row index is not saved

print(f"Results have been saved to {output_file}")

## Sort the list of predictive machine learning effects of the intermediate parameter VFA ##
results_df

def preprocess_data_ch4 (file_path):

    file_path = 'jz_data_knn.xlsx'
    data = pd.read_excel(file_path)

    X = data[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'TSf', 'VSf', 'HFV', 'HRT', 'Addition', 'Temperature', 'OLR' , 'TSo', 'VSo', 'pH', 'TAN', 'VFA']]
    y = data[['CH4']]     

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=0)

    scaler_X = StandardScaler()
    X_train = scaler_X.fit_transform(X_train)
    X_test = scaler_X.transform(X_test)

    scaler_y = StandardScaler()
    y_train = scaler_y.fit_transform(y_train.values.reshape(-1, 1)).flatten()
    y_test = scaler_y.transform(y_test.values.reshape(-1, 1)).flatten()

    return X_train, X_test, y_train, y_test, scaler_X, scaler_y

## Import data and train models ##
# Data import
X_train, X_test, y_train, y_test, scaler_X, scaler_y = preprocess_data_ch4 (r"D:\a\jz_data_knn.xlsx")

# train and evaluate models
results_df, trained_models, best_model = train_and_evaluate_model(X_train, y_train, X_test, y_test, scaler_y)

# Convert results_df to an Excel sheet and save it
output_file = 'model_evaluation_results_CH4.xlsx' 
results_df.to_excel(output_file, index=False)  

print(f"Results have been saved to {output_file}")

## Ranking of the prediction machine learning effect list of gas production performance parameter CH4 ##
results_df

def preprocess_data_Biogas_Production (file_path):
    file_path = 'jz_data_knn.xlsx'
    data = pd.read_excel(file_path)

    X = data[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'TSf', 'VSf', 'HFV', 'HRT', 'Addition', 'Temperature', 'OLR' , 'TSo', 'VSo', 'pH', 'TAN', 'VFA']]
    y = data[['Biogas_Production']]    

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=0)

    scaler_X = StandardScaler()
    X_train = scaler_X.fit_transform(X_train)
    X_test = scaler_X.transform(X_test)

    scaler_y = StandardScaler()
    y_train = scaler_y.fit_transform(y_train.values.reshape(-1, 1)).flatten()
    y_test = scaler_y.transform(y_test.values.reshape(-1, 1)).flatten()

    return X_train, X_test, y_train, y_test, scaler_X, scaler_y

## Import data and train models ##
# Data import
X_train, X_test, y_train, y_test, scaler_X, scaler_y = preprocess_data_Biogas_Production (r"D:\a\jz_data_knn.xlsx")

# train and evaluate models
results_df, trained_models, best_model = train_and_evaluate_model(X_train, y_train, X_test, y_test, scaler_y)

# Convert results_df to an Excel sheet and save it
output_file = 'model_evaluation_results_Biogas_Production.xlsx' 
results_df.to_excel(output_file, index=False)  

print(f"Results have been saved to {output_file}")

## Sorting of the list of predictive machine learning effects for gas production performance parameter Biogas_Production ##
results_df

# Shap analysis was performed on the prediction models of VFA, TAN, CH4 and BiogasProduction respectively, and the correlation between other variables and VFA, TAN, CH4 and BiogasProduction was judged#
### Part Three - Finding the correlation of other variables with VFA, TAN, CH4, Biogas_Production ###
# SHAP model explanation method
import shap
from matplotlib.transforms import ScaledTranslation 
import matplotlib.colors 
from matplotlib.cm import ScalarMappable
from matplotlib.colors import Normalize  

#######---------- TAN ------------#######

X_train, X_test, y_train, y_test, scaler_X, scaler_y = preprocess_data_tan (r"D:\a\jz_data_knn.xlsx")

# Hyperparameter optimization
param, best_score = optimizer_optuna(X_train, y_train, 30, "XGBoost")
filtered_param = {}
for key in param:
    if key in XGBRegressor().get_params().keys():
        filtered_param[key] = param[key]

model = XGBRegressor(**filtered_param)
model.fit(X_train, y_train)

explainer = shap.Explainer(model)
shap_values = explainer(X_train)

feature_names = [
    'PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV',
    'TSf', 'VSf', 'HFV', 'HRT', 'Addition', 'Temperature',
    'OLR', 'TSo', 'VSo', 'pH'
]

shap_values.feature_names = feature_names

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 8),
                              gridspec_kw={'width_ratios': [3, 1]},
                              sharey=True)  

mean_shap_values = np.abs(shap_values.values).mean(0)
feature_importance = pd.DataFrame(mean_shap_values, index=feature_names, columns=['importance'])
feature_importance = feature_importance.sort_values('importance', ascending=True)

plt.sca(ax1)
shap.plots.beeswarm(
    shap_values,
    show=False,
    max_display=len(feature_names),
    plot_size=(8, 8),
    color=plt.cm.plasma,  
    color_bar=False  
)

ax1.grid(True, alpha=0.3, color='gray', linestyle='--')

sorted_indices = np.argsort(np.abs(shap_values.values).mean(0))
sorted_feature_names = np.array(feature_names)[sorted_indices]

ax1.set_yticks(np.arange(len(sorted_feature_names)))
ax1.set_yticklabels(sorted_feature_names, fontsize=10, ha='right')

ax1.tick_params(axis='y', pad=5)

ax1.set_title('SHAP Value Beeswarm Plot', fontsize=14, fontweight='bold')

bars = ax2.barh(
    y=np.arange(len(feature_importance)),
    width=feature_importance['importance'],
    color='#D3D3D3',  
    edgecolor='black',  
    linewidth=1  
)

for spine in ax2.spines.values():
    spine.set_visible(False)

ax2.grid(False)

for bar in bars:
    width = bar.get_width()
    ax2.text(width, bar.get_y() + bar.get_height() / 2,
             '{:.3f}'.format(width),
             ha='left', va='center', fontsize=8)

ax2.set_title('Average |SHAP| value', fontsize=12, fontweight='bold')

ax2.set_yticklabels([])

dx = 0.1 / 2.54 / fig.dpi
offset = ScaledTranslation(dx, 0, fig.dpi_scale_trans)
for label in ax2.get_xticklabels():
    label.set_transform(label.get_transform() + offset)
ax2.set_position(ax2.get_position().translated(dx, 0))

for spine in ax1.spines.values():
    spine.set_visible(True)
    spine.set_color('black')
    spine.set_linewidth(2.0)

norm = Normalize(vmin=np.min(shap_values.values), vmax=np.max(shap_values.values))
sm = ScalarMappable(norm=norm, cmap=plt.cm.plasma)
sm.set_array([])

cbar_width = 0.1 / 2.54  # 宽度 0.1cm 转换为英寸
cbar_height = 0.5 / 2.54  # 高度 0.5cm 转换为英寸

ax2_pos = ax2.get_position()
cbar_x = ax2_pos.x1 - cbar_width
cbar_y = ax2_pos.y0
cbar_ax = fig.add_axes([cbar_x, cbar_y, cbar_width, cbar_height])

cbar = fig.colorbar(sm, cax=cbar_ax, orientation='vertical')
#cbar.set_ticks([0, 1])
cbar.set_ticks([norm.vmin, norm.vmax])
cbar.set_ticklabels(['Low', 'High'])
cbar.set_label('Feature Value', size=10)

plt.tight_layout()
plt.show()

#######---------- VFA ------------#######

X_train, X_test, y_train, y_test, scaler_X, scaler_y = preprocess_data_vfa (r"D:\a\jz_data_knn.xlsx")

# Hyperparameter optimization
param, best_score = optimizer_optuna(X_train, y_train, 30, "CatBoost")
filtered_param = {}
for key in param:
    if key in CatBoostRegressor().get_params().keys():
        filtered_param[key] = param[key]

model = CatBoostRegressor(**filtered_param)
model.fit(X_train, y_train)

explainer = shap.Explainer(model)
shap_values = explainer(X_train)

feature_names = [
    'PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV',
    'TSf', 'VSf', 'HFV', 'HRT', 'Addition', 'Temperature',
    'OLR', 'TSo', 'VSo', 'pH'
]

shap_values.feature_names = feature_names

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 8),
                              gridspec_kw={'width_ratios': [3, 1]},
                              sharey=True)  

mean_shap_values = np.abs(shap_values.values).mean(0)
feature_importance = pd.DataFrame(mean_shap_values, index=feature_names, columns=['importance'])
feature_importance = feature_importance.sort_values('importance', ascending=True)

plt.sca(ax1)
shap.plots.beeswarm(
    shap_values,
    show=False,
    max_display=len(feature_names),
    plot_size=(8, 8),
    color=plt.cm.plasma, 
    color_bar=False  
)

ax1.grid(True, alpha=0.3, color='gray', linestyle='--')

sorted_indices = np.argsort(np.abs(shap_values.values).mean(0))
sorted_feature_names = np.array(feature_names)[sorted_indices]

ax1.set_yticks(np.arange(len(sorted_feature_names)))
ax1.set_yticklabels(sorted_feature_names, fontsize=10, ha='right')

ax1.tick_params(axis='y', pad=5)

ax1.set_title('SHAP Value Beeswarm Plot', fontsize=14, fontweight='bold')

bars = ax2.barh(
    y=np.arange(len(feature_importance)),
    width=feature_importance['importance'],
    color='#D3D3D3', 
    edgecolor='black',  
    linewidth=1  
)

for spine in ax2.spines.values():
    spine.set_visible(False)

ax2.grid(False)

for bar in bars:
    width = bar.get_width()
    ax2.text(width, bar.get_y() + bar.get_height() / 2,
             '{:.3f}'.format(width),
             ha='left', va='center', fontsize=8)

ax2.set_title('Average |SHAP| value', fontsize=12, fontweight='bold')

ax2.set_yticklabels([])

dx = 0.1 / 2.54 / fig.dpi
offset = ScaledTranslation(dx, 0, fig.dpi_scale_trans)
for label in ax2.get_xticklabels():
    label.set_transform(label.get_transform() + offset)
ax2.set_position(ax2.get_position().translated(dx, 0))

for spine in ax1.spines.values():
    spine.set_visible(True)
    spine.set_color('black')
    spine.set_linewidth(2.0)

norm = Normalize(vmin=np.min(shap_values.values), vmax=np.max(shap_values.values))
sm = ScalarMappable(norm=norm, cmap=plt.cm.plasma)
sm.set_array([])

cbar_width = 0.1 / 2.54  # 宽度 0.1cm 转换为英寸
cbar_height = 0.5 / 2.54  # 高度 0.5cm 转换为英寸

ax2_pos = ax2.get_position()
cbar_x = ax2_pos.x1 - cbar_width
cbar_y = ax2_pos.y0
cbar_ax = fig.add_axes([cbar_x, cbar_y, cbar_width, cbar_height])

cbar = fig.colorbar(sm, cax=cbar_ax, orientation='vertical')
#cbar.set_ticks([0, 1])
cbar.set_ticks([norm.vmin, norm.vmax])
cbar.set_ticklabels(['Low', 'High'])
cbar.set_label('Feature Value', size=10)

plt.tight_layout()
plt.show()

#######---------- CH4 ------------#######

X_train, X_test, y_train, y_test, scaler_X, scaler_y = preprocess_data_ch4 (r"D:\a\jz_data_knn.xlsx")

# Hyperparameter optimization
param, best_score = optimizer_optuna(X_train, y_train, 30, "RandomForest")
filtered_param = {}
for key in param:
    if key in RandomForestRegressor().get_params().keys():
        filtered_param[key] = param[key]

model = RandomForestRegressor(**filtered_param)
model.fit(X_train, y_train)

explainer = shap.Explainer(model)
shap_values = explainer(X_train)

feature_names = [
    'PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 
    'TSf', 'VSf', 'HFV', 'HRT', 'Addition', 'Temperature', 
    'OLR', 'TSo', 'VSo', 'pH', 'TAN', 'VFA'
]

shap_values.feature_names = feature_names

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 8),
                              gridspec_kw={'width_ratios': [3, 1]},
                              sharey=True)  

mean_shap_values = np.abs(shap_values.values).mean(0)
feature_importance = pd.DataFrame(mean_shap_values, index=feature_names, columns=['importance'])
feature_importance = feature_importance.sort_values('importance', ascending=True)

plt.sca(ax1)
shap.plots.beeswarm(
    shap_values,
    show=False,
    max_display=len(feature_names),
    plot_size=(8, 8),
    color=plt.cm.plasma,  
    color_bar=False  
)

ax1.grid(True, alpha=0.3, color='gray', linestyle='--')

sorted_indices = np.argsort(np.abs(shap_values.values).mean(0))
sorted_feature_names = np.array(feature_names)[sorted_indices]

ax1.set_yticks(np.arange(len(sorted_feature_names)))
ax1.set_yticklabels(sorted_feature_names, fontsize=10, ha='right')

ax1.tick_params(axis='y', pad=5)

ax1.set_title('SHAP Value Beeswarm Plot', fontsize=14, fontweight='bold')

bars = ax2.barh(
    y=np.arange(len(feature_importance)),
    width=feature_importance['importance'],
    color='#D3D3D3',  
    edgecolor='black',  
    linewidth=1  
)

for spine in ax2.spines.values():
    spine.set_visible(False)

ax2.grid(False)

for bar in bars:
    width = bar.get_width()
    ax2.text(width, bar.get_y() + bar.get_height() / 2,
             '{:.3f}'.format(width),
             ha='left', va='center', fontsize=8)

ax2.set_title('Average |SHAP| value', fontsize=12, fontweight='bold')

ax2.set_yticklabels([])

dx = 0.1 / 2.54 / fig.dpi
offset = ScaledTranslation(dx, 0, fig.dpi_scale_trans)
for label in ax2.get_xticklabels():
    label.set_transform(label.get_transform() + offset)
ax2.set_position(ax2.get_position().translated(dx, 0))

for spine in ax1.spines.values():
    spine.set_visible(True)
    spine.set_color('black')
    spine.set_linewidth(2.0)

norm = Normalize(vmin=np.min(shap_values.values), vmax=np.max(shap_values.values))
sm = ScalarMappable(norm=norm, cmap=plt.cm.plasma)
sm.set_array([])

cbar_width = 0.1 / 2.54  # 宽度 0.1cm 转换为英寸
cbar_height = 0.5 / 2.54  # 高度 0.5cm 转换为英寸

ax2_pos = ax2.get_position()
cbar_x = ax2_pos.x1 - cbar_width
cbar_y = ax2_pos.y0
cbar_ax = fig.add_axes([cbar_x, cbar_y, cbar_width, cbar_height])

cbar = fig.colorbar(sm, cax=cbar_ax, orientation='vertical')
#cbar.set_ticks([0, 1])
cbar.set_ticks([norm.vmin, norm.vmax])
cbar.set_ticklabels(['Low', 'High'])
cbar.set_label('Feature Value', size=10)

plt.tight_layout()
plt.show()

#######---------- Biogas_Production ------------#######

X_train, X_test, y_train, y_test, scaler_X, scaler_y = preprocess_data_Biogas_Production (r"D:\a\jz_data_knn.xlsx")

# Hyperparameter optimization
param, best_score = optimizer_optuna(X_train, y_train, 30, "RandomForest")
filtered_param = {}
for key in param:
    if key in RandomForestRegressor().get_params().keys():
        filtered_param[key] = param[key]

model = RandomForestRegressor(**filtered_param)
model.fit(X_train, y_train)

explainer = shap.Explainer(model)
shap_values = explainer(X_train)

feature_names = [
    'PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 
    'TSf', 'VSf', 'HFV', 'HRT', 'Addition', 'Temperature', 
    'OLR', 'TSo', 'VSo', 'pH', 'TAN', 'VFA'
]

shap_values.feature_names = feature_names

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 8),
                              gridspec_kw={'width_ratios': [3, 1]},
                              sharey=True)  

mean_shap_values = np.abs(shap_values.values).mean(0)
feature_importance = pd.DataFrame(mean_shap_values, index=feature_names, columns=['importance'])
feature_importance = feature_importance.sort_values('importance', ascending=True)

plt.sca(ax1)
shap.plots.beeswarm(
    shap_values,
    show=False,
    max_display=len(feature_names),
    plot_size=(8, 8),
    color=plt.cm.plasma,  
    color_bar=False  
)

ax1.grid(True, alpha=0.3, color='gray', linestyle='--')

sorted_indices = np.argsort(np.abs(shap_values.values).mean(0))
sorted_feature_names = np.array(feature_names)[sorted_indices]

ax1.set_yticks(np.arange(len(sorted_feature_names)))
ax1.set_yticklabels(sorted_feature_names, fontsize=10, ha='right')

ax1.tick_params(axis='y', pad=5)

ax1.set_title('SHAP Value Beeswarm Plot', fontsize=14, fontweight='bold')

bars = ax2.barh(
    y=np.arange(len(feature_importance)),
    width=feature_importance['importance'],
    color='#D3D3D3',  
    edgecolor='black', 
    linewidth=1 
)

for spine in ax2.spines.values():
    spine.set_visible(False)

ax2.grid(False)

for bar in bars:
    width = bar.get_width()
    ax2.text(width, bar.get_y() + bar.get_height() / 2,
             '{:.3f}'.format(width),
             ha='left', va='center', fontsize=8)

ax2.set_title('Average |SHAP| value', fontsize=12, fontweight='bold')

ax2.set_yticklabels([])

dx = 0.1 / 2.54 / fig.dpi
offset = ScaledTranslation(dx, 0, fig.dpi_scale_trans)
for label in ax2.get_xticklabels():
    label.set_transform(label.get_transform() + offset)
ax2.set_position(ax2.get_position().translated(dx, 0))

for spine in ax1.spines.values():
    spine.set_visible(True)
    spine.set_color('black')
    spine.set_linewidth(2.0)

norm = Normalize(vmin=np.min(shap_values.values), vmax=np.max(shap_values.values))
sm = ScalarMappable(norm=norm, cmap=plt.cm.plasma)
sm.set_array([])

cbar_width = 0.1 / 2.54  # 宽度 0.1cm 转换为英寸
cbar_height = 0.5 / 2.54  # 高度 0.5cm 转换为英寸

ax2_pos = ax2.get_position()
cbar_x = ax2_pos.x1 - cbar_width
cbar_y = ax2_pos.y0
cbar_ax = fig.add_axes([cbar_x, cbar_y, cbar_width, cbar_height])

cbar = fig.colorbar(sm, cax=cbar_ax, orientation='vertical')
#cbar.set_ticks([0, 1])
cbar.set_ticks([norm.vmin, norm.vmax])
cbar.set_ticklabels(['Low', 'High'])
cbar.set_label('Feature Value', size=10)

plt.tight_layout()
plt.show()

# According to the optimal prediction models of VFA, TAN, CH4 and Biogas_Production above, the causal inference DML machine learning was introduced to calculate the ATE #
### Part Four - The causal relationships of VFA-CH4, VFA-BiogasProduction, TAN-CH4 and TAN-BiogasProduction were obtained ###
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from econml.dml import DML
# Cross-validation and hyperparameter optimization
from sklearn.model_selection import train_test_split  # Splitting dataset
# Functions for model performance evaluation metrics
import seaborn as sns
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from xgboost import XGBRegressor
from catboost import CatBoostRegressor
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows  
plt.rcParams['font.sans-serif'] =['Times New Roman']

# ######-----------TAN to CH4-------------###### #
def causal(X, Y, T, W, output_excel="TAN to CH4.xlsx"):
    '''Causal Inference Main Function: Evaluate the causal effect of the processing variable (T) on the outcome variable (Y) using double machine learning (DML).
    parameter:
        X (pd.DataFrame): Covariates (contextual features)
        Y (pd.Series): Outcome Variable (Target Indicator CH4)
        T (pd.Series): Handling Variables (Feature of Interest TAN)
        W (pd.DataFrame): Other control variables
        output_excel (str): The name of the resulting output file
    '''

    # 数据分割逻辑 ===============================================
    '''Data Segmentation Strategy (Random Segmentation or Random Segmentation of Groups)'''
    if (X == 0).all().all(): 
        # When no valid covariates are present, only Y, T, W are segmented
        Y_train, Y_test, T_train, T_test, W_train, W_test = train_test_split(
            Y, T, W, test_size=0.2, random_state=0)
        X_train_combined = W_train  
    else:
        # The usual case includes the segmentation of the covariate X
        Y_train, Y_test, T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(
            Y, T, X, W, test_size=0.2, random_state=0)
        X_train_combined = np.hstack((X_train, W_train))  

    param_Y = {}  
    param_T = {}  

    param_Y, best_score_Y = optimizer_optuna(
        X_train_combined, Y_train, 
        n_trials=30, 
        model_type="RandomForest"  
    )

    param_T, best_score_T = optimizer_optuna(
        X_train_combined, T_train,
        n_trials=30,
        model_type="XGBoost"
    )

    # Build a DML model 
    model = DML(
        model_y=RandomForestRegressor(**param_Y),  # Phase 1 Y Prediction Model
        model_t=XGBRegressor(**param_T),          # Phase 1 T Prediction Model
        model_final=LinearRegression()            # Final Effect Estimation Model
    )
    
    # Model training and inference 
    '''Select the training mode based on the data scenario'''
    if (X == 0).all().all():  
        model.fit(Y_train, T_train, 
                 X=None,        
                 W=W_train,      
                 inference="bootstrap") 
        
        treatment_effects = np.zeros((T.shape[0], 1))  
        effect_values = model.effect()  
        treatment_effects[:] = effect_values.reshape(-1, 1)  
        
        inference_result = None
        lb, ub = None, None  
        p_value = None      
    
    else:  
        model.fit(Y_train, T_train, 
                 X=X_train, 
                 W=W_train, 
                 inference="bootstrap")
        
        treatment_effects = model.effect(X)  
        inference_result = model.const_marginal_effect_inference(X).summary_frame()
        lb, ub = model.const_marginal_effect_interval(X, alpha=0.05)
        
        p_value = inference_result["p_value"] if "p_value" in inference_result.columns else None

    # Visual analytics
    '''处理效应分布直方图'''
    plt.figure(figsize=(6, 6))
    sns.histplot(treatment_effects, kde=True, bins=30)
    plt.title('Estimated Treatment Effects Distribution')
    plt.xlabel('Causal Effect Magnitude')
    plt.ylabel('Frequency Count')
    plt.show()

    '''Scatter plot of effect value vs. eigenvalue'''
    plt.figure(figsize=(6, 6))
    plt.scatter(T, treatment_effects, alpha=0.5)
    plt.title('Treatment Effect vs Feature Value')
    plt.xlabel('TAN Concentration Level')  
    plt.ylabel('Estimated Causal Impact')
    plt.show()

    # Statistical analysis of results
    '''Calculate key statistical indicators'''
    ate = np.mean(treatment_effects)  
    confidence_interval = (np.mean(lb), np.mean(ub)) if lb.size > 0 and ub.size > 0 else None
    avg_p = p_value.mean() if p_value is not None else None

    # The console outputs key results
    print(f"ATE: {ate:.4f}")
    if confidence_interval:
        print(f"95% confidence interval: [{confidence_interval[0]:.4f}, {confidence_interval[1]:.4f}]")
    print(f"Average P_value: {avg_p if avg_p else 'N/A'}")

    # Persistence of results
    wb = Workbook()
    ws = wb.active
    ws.title = "Causal Analysis Report"

    # Write metadata
    ws.append(["关键指标", "值"])
    ws.append(["ATE", ate])
    ws.append(["置信区间下限", confidence_interval[0] if confidence_interval else "N/A"])
    ws.append(["置信区间上限", confidence_interval[1] if confidence_interval else "N/A"])
    ws.append(["平均P值", avg_p if avg_p else "N/A"])

    # Create a detailed results table
    ws = wb.create_sheet("Detailed Effects")
    treatment_effects_df = pd.DataFrame({
        "样本ID": X.index if hasattr(X, 'index') else range(len(treatment_effects)),
        "处理效应值": treatment_effects.flatten(),
        "P值": p_value if p_value is not None else ["N/A"]*len(treatment_effects)
    })
    
    # Write detailed data
    for r in dataframe_to_rows(treatment_effects_df, index=False, header=True):
        ws.append(r)

    # Save the file
    wb.save(output_excel)
    print(f"The results of the analysis have been saved to: {output_excel}")

    return

# Read the data and call the causal inference function
file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['CH4']]      ##### Final gas production performance parameters 'CH4' , 'Biogas_Production'
T = dataset[['TAN']]      ##### Intermediate parameters 'TAN', 'VFA'
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'HFV', 'Addition', 'Temperature', 'OLR', 'HRT', 'pH']]
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]
causal(X, Y, T, W, output_excel="TAN to CH4.xlsx")

# ######-----------VFA to CH4-------------###### #
def causal(X, Y, T, W, output_excel="VFA to CH4.xlsx"):
    '''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
    if (X == 0).all().all():
        Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
        X_train_combined =W_train
    else:
        print(1)
        Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
        X_train_combined = np.hstack((X_train, W_train))
    '''Hyperparameter optimization'''
    param_Y = {}
    param_T = {}
    '''Hyperparameter optimization'''
    param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
    param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "CatBoost")

    # Build a causal inference model
    model = DML(model_y=RandomForestRegressor(**param_Y),
                model_t=CatBoostRegressor(**param_T),
                model_final=LinearRegression())
    
    # Train the model and enable bootstrap inference
    if (X == 0).all().all():
        model.fit(Y_train, T_train, X=None, W=W_train, inference="bootstrap")
        treatment_effects = np.zeros((T.shape[0], 1))
        effect_values = model.effect()
        treatment_effects[:] = effect_values.reshape(-1, 1)
        inference_result = None
        lb, ub = None, None
        p_value = None
    else:
        model.fit(Y_train, T_train, X=X_train, W=W_train, inference="bootstrap")
        treatment_effects = model.effect(X)
        inference_result = model.const_marginal_effect_inference(X).summary_frame()
        lb, ub = model.const_marginal_effect_interval(X, alpha=0.05)

        # Check if there is a p_value column
        if "p_value" in inference_result.columns:
            p_value = inference_result["p_value"]
        else:
            p_value = None


    # Draw a histogram of causal effects
    plt.figure(figsize=(6, 6))
    sns.histplot(treatment_effects, kde=True, bins=30)
    plt.title('Distribution of Estimated Treatment Effects')
    plt.xlabel('Estimated Causal Effect')
    plt.ylabel('Frequency')
    plt.show()
    
    plt.figure(figsize=(6, 6))
    plt.scatter(T, treatment_effects, alpha=0.5)
    plt.title('Causal Effect')
    plt.xlabel('Feature Value')
    plt.ylabel('Estimated Causal Effect')
    plt.show()

    # Output ATE, confidence interval, and p-value
    ate = np.mean(treatment_effects)
    confidence_interval = (np.mean(lb), np.mean(ub)) if lb is not None and ub is not None else None
    p_val = p_value.mean() if p_value is not None else "Not available"
    
    print(f"ATE (Average Treatment Effect)：{ate}")
    if confidence_interval:
        print(f"95% Confidence Interval for ATE: [{confidence_interval[0]}, {confidence_interval[1]}]")
    print(f"P-value: {p_val}")

    # Save the results to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Causal Results"

    # Write ATE, confidence interval and p-value
    ws.append(["ATE", "95% Confidence Interval Lower", "95% Confidence Interval Upper", "P-value"])
    ws.append([ate, confidence_interval[0] if confidence_interval else "N/A", confidence_interval[1] if confidence_interval else "N/A", p_value.mean() if p_value is not None else "Not available"])

    # Write causal effect results to Excel
    treatment_effects_df = pd.DataFrame({
        "Estimated Causal Effect": treatment_effects.flatten(),
        "P-value": p_value if p_value is not None else ["Not available"] * len(treatment_effects)
    })

    for r in dataframe_to_rows(treatment_effects_df, index=False, header=True):
        ws.append(r)

    # Save the Excel file
    wb.save(output_excel)

    print(f"Results and figures have been saved to {output_excel}.")

    return

# Read the data and call the causal inference function
file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['CH4']]      ##### 'CH4' , 'Biogas_Production'
T = dataset[['VFA']]       ##### 'TAN', 'VFA'
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'HFV', 'Addition', 'Temperature', 'OLR', 'HRT', 'pH']]
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]
causal(X, Y, T, W, output_excel="VFA to CH4.xlsx")

# ######-----------TAN to Biogas_Production-------------###### #
def causal(X, Y, T, W, output_excel="TAN to Biogas_Production.xlsx"):
    '''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
    if (X == 0).all().all():
        Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
        X_train_combined =W_train
    else:
        print(1)
        Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
        X_train_combined = np.hstack((X_train, W_train))
    '''Hyperparameter optimization'''
    param_Y = {}
    param_T = {}
    '''Hyperparameter optimization'''
    param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
    param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "XGBoost")

    model = DML(model_y=RandomForestRegressor(**param_Y),
                model_t=XGBRegressor(**param_T),
                model_final=LinearRegression())
    
    if (X == 0).all().all():
        model.fit(Y_train, T_train, X=None, W=W_train, inference="bootstrap")
        treatment_effects = np.zeros((T.shape[0], 1))
        effect_values = model.effect()
        treatment_effects[:] = effect_values.reshape(-1, 1)
        lb, ub = None, None
        p_value = None
    else:
        model.fit(Y_train, T_train, X=X_train, W=W_train, inference="bootstrap")
        treatment_effects = model.effect(X)
        lb, ub = model.const_marginal_effect_interval(X, alpha=0.05)

        inference_result = model.const_marginal_effect_inference(X).summary_frame()
        print("Inference summary frame:")
        print(inference_result.head())  

        if "p_value" in inference_result.columns:
            p_value = inference_result["p_value"]
        else:
            p_value = None

    plt.figure(figsize=(6, 6))
    sns.histplot(treatment_effects, kde=True, bins=30)
    plt.title('Distribution of Estimated Treatment Effects')
    plt.xlabel('Estimated Causal Effect')
    plt.ylabel('Frequency')
    plt.show()
    
    plt.figure(figsize=(6, 6))
    plt.scatter(T, treatment_effects, alpha=0.5)
    plt.title('Causal Effect')
    plt.xlabel('Feature Value')
    plt.ylabel('Estimated Causal Effect')
    plt.show()

    ate = np.mean(treatment_effects)
    confidence_interval = (np.mean(lb), np.mean(ub)) if lb is not None and ub is not None else None
    p_val = p_value.mean() if p_value is not None else "Not available"
    
    print(f"ATE (Average Treatment Effect)：{ate}")
    if confidence_interval:
        print(f"95% Confidence Interval for ATE: [{confidence_interval[0]}, {confidence_interval[1]}]")
    print(f"P-value: {p_val}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Causal Results"
    
    ws.append(["ATE", "95% Confidence Interval Lower", "95% Confidence Interval Upper", "P-value"])
    ws.append([ate, confidence_interval[0] if confidence_interval else "N/A", confidence_interval[1] if confidence_interval else "N/A", p_val])
    
    treatment_effects_df = pd.DataFrame(treatment_effects, columns=["Estimated Causal Effect"])
    for r in dataframe_to_rows(treatment_effects_df, index=False, header=True):
        ws.append(r)
    
    wb.save(output_excel)

    print(f"Results and figures have been saved to {output_excel}.")

    return

file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['Biogas_Production']]      ##### 'CH4' , 'Biogas_Production'
T = dataset[['TAN']]       ##### 'TAN', 'VFA'
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'HFV', 'Addition', 'Temperature', 'OLR', 'HRT', 'pH']]
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]
causal(X, Y, T, W, output_excel="TAN to Biogas_Production.xlsx")

# ######-----------VFA to Biogas_Production-------------###### #
def causal(X, Y, T, W, output_excel="VFA to Biogas_Production.xlsx"):
    '''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
    if (X == 0).all().all():
        Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
        X_train_combined =W_train
    else:
        print(1)
        Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
        X_train_combined = np.hstack((X_train, W_train))
    '''Hyperparameter optimization'''
    param_Y = {}
    param_T = {}
    '''Hyperparameter optimization'''
    param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
    param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "CatBoost")

    model = DML(model_y=RandomForestRegressor(**param_Y),
                model_t=CatBoostRegressor(**param_T),
                model_final=LinearRegression())
    
    if (X == 0).all().all():
        model.fit(Y_train, T_train, X=None, W=W_train, inference="bootstrap")
        treatment_effects = np.zeros((T.shape[0], 1))
        effect_values = model.effect()
        treatment_effects[:] = effect_values.reshape(-1, 1)
        lb, ub = None, None
        p_value = None
    else:
        model.fit(Y_train, T_train, X=X_train, W=W_train, inference="bootstrap")
        treatment_effects = model.effect(X)
        lb, ub = model.const_marginal_effect_interval(X, alpha=0.05)

        inference_result = model.const_marginal_effect_inference(X).summary_frame()
        print("Inference summary frame:")
        print(inference_result.head()) 

        if "p_value" in inference_result.columns:
            p_value = inference_result["p_value"]
        else:
            p_value = None

    plt.figure(figsize=(6, 6))
    sns.histplot(treatment_effects, kde=True, bins=30)
    plt.title('Distribution of Estimated Treatment Effects')
    plt.xlabel('Estimated Causal Effect')
    plt.ylabel('Frequency')
    plt.show()

    plt.figure(figsize=(6, 6))
    plt.scatter(T, treatment_effects, alpha=0.5)
    plt.title('Causal Effect')
    plt.xlabel('Feature Value')
    plt.ylabel('Estimated Causal Effect')
    plt.show()

    ate = np.mean(treatment_effects)
    confidence_interval = (np.mean(lb), np.mean(ub)) if lb is not None and ub is not None else None
    p_val = p_value.mean() if p_value is not None else "Not available"
    
    print(f"ATE (Average Treatment Effect)：{ate}")
    if confidence_interval:
        print(f"95% Confidence Interval for ATE: [{confidence_interval[0]}, {confidence_interval[1]}]")
    print(f"P-value: {p_val}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Causal Results"
    
    ws.append(["ATE", "95% Confidence Interval Lower", "95% Confidence Interval Upper", "P-value"])
    ws.append([ate, confidence_interval[0] if confidence_interval else "N/A", confidence_interval[1] if confidence_interval else "N/A", p_val])
    
    treatment_effects_df = pd.DataFrame(treatment_effects, columns=["Estimated Causal Effect"])
    for r in dataframe_to_rows(treatment_effects_df, index=False, header=True):
        ws.append(r)
    
    wb.save(output_excel)

    print(f"Results and figures have been saved to {output_excel}.")

    return

file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['Biogas_Production']]      ##### 'CH4' , 'Biogas_Production'
T = dataset[['VFA']]       ##### 'TAN', 'VFA'
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'HFV', 'Addition', 'Temperature', 'OLR', 'HRT', 'pH']]
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]
causal(X, Y, T, W, output_excel="VFA to Biogas_Production.xlsx")

### Part Five - Causal analysis machine learning is used to explore the causal effects between multiple variables and multivariates, and to construct causal networks ###
from econml.solutions.causal_analysis import CausalAnalysis
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import networkx as nx
from sklearn.preprocessing import StandardScaler
## Determine causal relationships between other variables and intermediate parameters ##
file_path = 'jz_data_knn.xlsx'
data = pd.read_excel(file_path)

features = data[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'TSf', 'VSf', 'HFV', 'HRT', 'Addition', 'Temperature', 'OLR' , 'TSo', 'VSo', 'pH']]

########----------- VFA -----------########

label = data['VFA']

scaler = StandardScaler()
features_scaled = pd.DataFrame(scaler.fit_transform(features), columns=features.columns)

ca = CausalAnalysis(
    feature_inds=features.columns,
    categorical=[],
    classification=False,
    nuisance_models="automl",
    heterogeneity_model="linear",
    random_state=123,
    n_jobs=-1
)
ca.fit(features_scaled, np.ravel(label))

global_causal_effects = ca.global_causal_effect(alpha=0.05)
global_causal_effects_sorted = global_causal_effects.sort_values(by="p_value")

with pd.ExcelWriter('Causal_relationship_VFA.xlsx') as writer:
    global_causal_effects_sorted.to_excel(writer, sheet_name='Causal Effects')


########----------- TAN -----------########
label = data['TAN']

scaler = StandardScaler()
features_scaled = pd.DataFrame(scaler.fit_transform(features), columns=features.columns)

ca = CausalAnalysis(
    feature_inds=features.columns,
    categorical=[],
    classification=False,
    nuisance_models="automl",
    heterogeneity_model="linear",
    random_state=123,
    n_jobs=-1
)
ca.fit(features_scaled, np.ravel(label))

global_causal_effects = ca.global_causal_effect(alpha=0.05)
global_causal_effects_sorted = global_causal_effects.sort_values(by="p_value")

with pd.ExcelWriter('Causal_relationship_TAN.xlsx') as writer:
    global_causal_effects_sorted.to_excel(writer, sheet_name='Causal Effects')


## Determine causal relationships between other variables and gas production performance parameters ##
# Define feature and label variables
# Select a field with 19 characteristic columns (specific fields related to biogas engineering parameters)
features = data[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'TSf', 'VSf', 'HFV', 'HRT', 'Addition', 'Temperature', 'OLR' , 'TSo', 'VSo', 'VFA', 'pH' , 'TAN']]
########----------- CH4 -----------########
label = data['CH4']  # Select CH4 column as the target variable (methane yield)

# Data standardization preprocessing
scaler = StandardScaler()  # Create a normalizer instance (mean of 0 and standard deviation of 1)
features_scaled = pd.DataFrame(scaler.fit_transform(features),  
                               columns=features.columns)  

# Initialize the causal analysis model
ca = CausalAnalysis(
    feature_inds=features.columns,
    categorical=[],  
    classification=False,  
    nuisance_models="automl",  
    heterogeneity_model="linear",  
    random_state=123,  
    n_jobs=-1  
)
# Training a causal analysis model (passing in normalized features and one-dimensional label data)
ca.fit(features_scaled, np.ravel(label))  # np.ravel() Make sure the label is a one-dimensional array

# Calculate and rank global causal effects
global_causal_effects = ca.global_causal_effect(alpha=0.05)  # Calculate causal effects with 95% confidence
global_causal_effects_sorted = global_causal_effects.sort_values(by="p_value")  # Sort by p-value in ascending order

# Write the results of causal analysis to an Excel file
with pd.ExcelWriter('Causal_relationship_CH4.xlsx') as writer:  
    global_causal_effects_sorted.to_excel(writer, sheet_name='Causal Effects')


########----------- Biogas_Production -----------########

label = data['Biogas_Production']  

scaler = StandardScaler()  
features_scaled = pd.DataFrame(scaler.fit_transform(features),  
                               columns=features.columns)  

ca = CausalAnalysis(
    feature_inds=features.columns,  
    categorical=[],  
    classification=False,  
    nuisance_models="automl",  
    heterogeneity_model="linear",  
    random_state=123,  
    n_jobs=-1  
)

ca.fit(features_scaled, np.ravel(label))  

global_causal_effects = ca.global_causal_effect(alpha=0.05) 
global_causal_effects_sorted = global_causal_effects.sort_values(by="p_value")  

with pd.ExcelWriter('Causal_relationship_Biogas_Production.xlsx') as writer:  
    global_causal_effects_sorted.to_excel(writer,  
                                         sheet_name='Causal Effects')  

from econml.solutions.causal_analysis import CausalAnalysis
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import networkx as nx
from sklearn.preprocessing import StandardScaler
## Determine the causal relationship between all characteristic variables ##
file_path = 'jz_data_knn.xlsx'
data = pd.read_excel(file_path)

# Define features and labels
features = data[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'TSf', 'VSf', 'HFV', 'HRT', 'Addition', 'Temperature', 'OLR', 'VFA', 'pH', 'TSo', 'VSo', 'TAN']]

scaler = StandardScaler()
features_scaled = pd.DataFrame(scaler.fit_transform(features), columns=features.columns)

l_names = ['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'TSf', 'VSf', 'HFV', 'HRT', 'Addition', 'Temperature', 'OLR', 'VFA', 'pH', 'TSo', 'VSo', 'TAN']

result = pd.DataFrame()
for l_name in l_names:
    x = features_scaled.drop(columns=[l_name])
    y = pd.DataFrame(features_scaled[l_name])
    
    ca = CausalAnalysis(
        feature_inds=x.columns,
        categorical=[],
        classification=False,
        nuisance_models="automl",
        heterogeneity_model="linear",
        n_jobs=-1,
        random_state=123)
    ca.fit(x, np.ravel(y))

    global_summ = ca.global_causal_effect(alpha=0.05)
    global_summ_sorted = global_summ.sort_values(by="p_value")
    
    global_summ.insert(0, 'Function', l_name)
    global_summ.insert(0, 'label_feature', y.columns[0])
    
    result = pd.concat([result, global_summ])
    
with pd.ExcelWriter('Causal_relationship_among_features.xlsx') as writer:
    result.to_excel(writer)


### Part Six - Dowhy is used to construct a counterfactual prediction model based on the DML causal inference model to realize engineering prediction at the algorithm level ###
from sklearn.preprocessing import PolynomialFeatures
from sklearn.model_selection import train_test_split   # Splitting dataset
import numpy as np
import pandas as pd
import dowhy.do_samplers as do_samplers
from dowhy.causal_model import CausalModel
# from dowhy.utils.api import parse_state11
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from xgboost import XGBRegressor
from catboost import CatBoostRegressor
from sklearn.linear_model import LassoCV

def do(
        self,
        x,
        method="weighting",
        num_cores=1,
        variable_types={},
        outcome=None,
        params=None,
        dot_graph=None,
        common_causes=None,
        estimand_type="nonparametric-ate",
        proceed_when_unidentifiable=False,
        stateful=False,
    ):
    x, keep_original_treatment = self.parse_x(x)
    outcome = parse_state(outcome)
    if not stateful or method != self._method:
            self.reset()
    if not self._causal_model:
        self._causal_model = CausalModel(
            self._obj,
            [xi for xi in x.keys()],
            outcome,
            graph=dot_graph,
            common_causes=common_causes,
            instruments=None,
            estimand_type=estimand_type,
            proceed_when_unidentifiable=proceed_when_unidentifiable,
        )
    # self._identified_estimand = self._causal_model.identify_effect()

    if not bool(variable_types):  # check if the variables dictionary is empty
        variable_types = dict(self._obj.dtypes)  # Convert the series containing data types to a dictionary
        for key in variable_types.keys():
            variable_types[key] = self.convert_to_custom_type(
                variable_types[key].name
            )  # Obtain the custom type corrosponding to each data type

    elif len(self._obj.columns) > len(variable_types):
        all_variables = dict(self._obj.dtypes)
        for key in all_variables.keys():
            if key not in variable_types:
                variable_types[key] = self.convert_to_custom_type(all_variables[key].name)

    elif len(self._obj.columns) < len(variable_types):
        raise Exception("Number of variables in the DataFrame is lesser than the variable_types dict")

    if not self._sampler:
        self._method = method
        do_sampler_class = do_samplers.get_class_object(method + "_sampler")
        self._sampler = do_sampler_class(
            self._obj,
            # self._identified_estimand,
            # self._causal_model._treatment,
            # self._causal_model._outcome,
            params=params,
            variable_types=variable_types,
            num_cores=num_cores,
            causal_model=self._causal_model,
            keep_original_treatment=keep_original_treatment,
        )
    result = self._sampler.do_sample(x)
    if not stateful:
        self.reset()
    return result

######----------TAN-----------######

import dowhy
from dowhy import CausalModel
import pandas as pd

file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['Biogas_Production']]  
T = dataset[['TAN']]  
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'HFV', 'Addition', 'Temperature', 'OLR', 'HRT', 'pH', 'VFA']]  
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]  

causal_graph = """
digraph {
    TAN -> Biogas_Production;
    PWFV -> TAN;
    FW -> TAN;
    ICO -> TAN;
    AO -> TAN;
    Manure -> TAN;
    Percolate -> TAN;
    Temperature -> TAN;
    OLR -> TAN;
    HRT -> TAN;
    pH -> TAN;
    VFA -> TAN;
    TSf -> TAN;
    VSf -> TAN;
    TSo -> TAN;
    VSo -> TAN;
    TSf -> Biogas_Production;
    VSf -> Biogas_Production;
    TSo -> Biogas_Production;
    VSo -> Biogas_Production;
}
"""

# I. Create a causal model from the data and domain knowledge.
model = CausalModel(
    data=dataset,
    treatment='TAN',
    outcome='Biogas_Production',
    common_causes=['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'HFV', 'Addition', 'Temperature', 'OLR', 'HRT', 'pH', 'VFA', 'TSf', 'VSf', 'TSo', 'VSo'],
    graph=causal_graph
)

model.view_model()
# layout="dot"
from IPython.display import Image, display
display(Image(filename="causal_model.png"))

# II. Identify causal effect and return target estimands
identified_estimand = model.identify_effect()
print(identified_estimand)

# III. Estimate the target estimand using a statistical method.
'''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
if (X == 0).all().all():
    Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
    X_train_combined =W_train
else:
    print(1)
    Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
    X_train_combined = np.hstack((X_train, W_train))
'''Hyperparameter optimization'''
param_Y = {}
param_T = {}
'''Hyperparameter optimization'''
param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "XGBoost")

### Calculate the baseline biogas production (TAN = 0)   ###
baseline_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=0,  # TAN = 0, the baseline situation
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  "model_t": XGBRegressor(**param_T),
                                  "model_final": LassoCV(),
                                  "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                 }}
)

baseline_biogas_production = baseline_estimate.value  # An estimate of the baseline yield is obtained
print("Baseline Biogas Production (TAN = 0):", baseline_biogas_production)

dml_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=1,
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  'model_t': XGBRegressor(**param_T),
                                  "model_final":LassoCV(),
                                  'featurizer':PolynomialFeatures(degree=2, include_bias=True)
                                               }}
)

print("DML ATE Estimate:", dml_estimate)

# IV. Refute the obtained estimate using multiple robustness checks.
refute_results = model.refute_estimate(identified_estimand, dml_estimate,
                                       method_name="placebo_treatment_refuter")
print(refute_results)

# Iterate over the range of TAN from 0 to 4000 and perform intervention analysis
# Traverse the extent of the TAN and perform an intervention analysis to calculate the final biogas production forecast   #############
results = []
for tan_value in range(0, 4001, 1):
    # Estimate causal effects for each TAN value
    estimate = model.estimate_effect(
        identified_estimand,
        method_name="backdoor.econml.dml.DML",
        control_value=0,
        treatment_value=tan_value,
        method_params={"init_params":{'model_y': RandomForestRegressor(**param_Y),
                                      "model_t": XGBRegressor(**param_T),  
                                      "model_final": LassoCV(),  
                                      "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                     }}
    )
    
    # Causal effects of TAN on biogas production
    causal_effect = estimate.value
    
    # By superimposing the baseline yield and causal effects, the final predicted biogas yield is obtained
    predicted_biogas_production_tan = baseline_biogas_production + causal_effect
    
    # Save the results
    results.append({
        "TAN": tan_value,
        "Predicted_Biogas_Production": predicted_biogas_production_tan
    })

# Convert the results into a DataFrame and save it as an Excel file
results_df = pd.DataFrame(results)
print(results_df)
results_df.to_excel("TAN_Predictions.xlsx", index=False)

print("直接预测的沼气产量结果已保存为 Excel 文件。")

######----------Addition-----------######

import dowhy
from dowhy import CausalModel
import pandas as pd

file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['Biogas_Production']]  
T = dataset[['Addition']]  
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'HRT', 'Temperature', 'OLR', 'HFV', 'pH', 'TAN']] 
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]  

causal_graph = """
digraph {
    Addition -> Biogas_Production;
    PWFV -> Addition;
    FW -> Addition;
    ICO -> Addition;
    AO -> Addition;
    Manure -> Addition;
    Percolate -> Addition;
    Temperature -> Addition;
    OLR -> Addition;
    HFV -> Addition;
    pH -> Addition;
    TAN -> Addition;
    TSf -> Addition;
    VSf -> Addition;
    TSo -> Addition;
    VSo -> Addition;
    TSf -> Biogas_Production;
    VSf -> Biogas_Production;
    TSo -> Biogas_Production;
    VSo -> Biogas_Production;
}
"""
# I. Create a causal model from the data and domain knowledge.
model = CausalModel(
    data=dataset,
    treatment='Addition',
    outcome='Biogas_Production',
    common_causes=['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'HRT', 'Temperature', 'OLR', 'HFV', 'pH', 'TAN', 'TSf', 'VSf', 'TSo', 'VSo'],
    graph=causal_graph
)

model.view_model()
# layout="dot"
from IPython.display import Image, display
display(Image(filename="causal_model.png"))

# II. Identify causal effect and return target estimands
identified_estimand = model.identify_effect()
print(identified_estimand)

# III. Estimate the target estimand using a statistical method.
'''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
if (X == 0).all().all():
    Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
    X_train_combined =W_train
else:
    print(1)
    Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
    X_train_combined = np.hstack((X_train, W_train))
'''Hyperparameter optimization'''
param_Y = {}
param_T = {}
'''Hyperparameter optimization'''
param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "RandomForest")

### Calculate the baseline biogas production (Addition = 0)   ###
baseline_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=0,  
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  "model_t": RandomForestRegressor(**param_T),
                                  "model_final": LassoCV(),
                                  "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                 }}
)

baseline_biogas_production = baseline_estimate.value  
print("Baseline Biogas Production (Addition = 0):", baseline_biogas_production)

dml_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=1,
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  'model_t': RandomForestRegressor(**param_T),
                                  "model_final":LassoCV(),
                                  'featurizer':PolynomialFeatures(degree=2, include_bias=True)
                                               }}
)

print("DML ATE Estimate:", dml_estimate)

# IV. Refute the obtained estimate using multiple robustness checks.
refute_results = model.refute_estimate(identified_estimand, dml_estimate,
                                       method_name="placebo_treatment_refuter")
print(refute_results)

# Iterate over the range of Addition from 0 to 900 and perform intervention analysis
results = []
for addition_value in np.arange(0, 126, 0.1):  
    estimate = model.estimate_effect(
        identified_estimand,
        method_name="backdoor.econml.dml.DML",
        control_value=0,
        treatment_value=[addition_value],
        method_params={"init_params":{'model_y': RandomForestRegressor(**param_Y),
                                      "model_t": RandomForestRegressor(**param_T),  
                                      "model_final": LassoCV(),  
                                      "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                     }}
    )

    causal_effect = estimate.value
    
    predicted_biogas_production_addition = baseline_biogas_production + causal_effect

    results.append({
        "Addition": addition_value,
        "Predicted_Biogas_Production": predicted_biogas_production_addition 
    })

results_df = pd.DataFrame(results)
print(results_df)
results_df.to_excel("Addition_Predictions.xlsx", index=False)

print("反事实预测结果已保存为 Excel 文件。")

######----------HFV-----------######

import dowhy
from dowhy import CausalModel
import pandas as pd

file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['Biogas_Production']]  
T = dataset[['HFV']]  
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'Addition', 'Temperature', 'OLR', 'HRT', 'pH', 'TAN']]  
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]  

causal_graph = """
digraph {
    HFV -> Biogas_Production;
    PWFV -> HFV;
    FW -> HFV;
    ICO -> HFV;
    AO -> HFV;
    Manure -> HFV;
    Percolate -> HFV;
    Temperature -> HFV;
    OLR -> HFV;
    HRT -> HFV;
    pH -> HFV;
    TAN -> HFV;
    TSf -> HFV;
    VSf -> HFV;
    TSo -> HFV;
    VSo -> HFV;
    TSf -> Biogas_Production;
    VSf -> Biogas_Production;
    TSo -> Biogas_Production;
    VSo -> Biogas_Production;
}
"""
# I. Create a causal model from the data and domain knowledge.
model = CausalModel(
    data=dataset,
    treatment='HFV',
    outcome='Biogas_Production',
    common_causes=['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'Addition', 'Temperature', 'OLR', 'HRT', 'pH', 'TAN', 'TSf', 'VSf', 'TSo', 'VSo'],
    graph=causal_graph
)

model.view_model()
# layout="dot"
from IPython.display import Image, display
display(Image(filename="causal_model.png"))

# II. Identify causal effect and return target estimands
identified_estimand = model.identify_effect()
print(identified_estimand)

# III. Estimate the target estimand using a statistical method.
'''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
if (X == 0).all().all():
    Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
    X_train_combined =W_train
else:
    print(1)
    Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
    X_train_combined = np.hstack((X_train, W_train))
'''Hyperparameter optimization'''
param_Y = {}
param_T = {}
'''Hyperparameter optimization'''
param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "RandomForest")

### Calculate the baseline biogas production (HFV = 0)   ###
baseline_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=0,  
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  "model_t": RandomForestRegressor(**param_T),
                                  "model_final": LassoCV(),
                                  "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                 }}
)

baseline_biogas_production = baseline_estimate.value  
print("Baseline Biogas Production (HFV = 0):", baseline_biogas_production)

dml_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=1,
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  'model_t': RandomForestRegressor(**param_T),
                                  "model_final":LassoCV(),
                                  'featurizer':PolynomialFeatures(degree=2, include_bias=True)
                                               }}
)

print("DML ATE Estimate:", dml_estimate)

# IV. Refute the obtained estimate using multiple robustness checks.
refute_results = model.refute_estimate(identified_estimand, dml_estimate,
                                       method_name="placebo_treatment_refuter")
print(refute_results)

# Iterate over the range of HFV from 0 to 380 and perform intervention analysis
results = []
for hfv_value in range(0, 381, 1):  
    estimate = model.estimate_effect(
        identified_estimand,
        method_name="backdoor.econml.dml.DML",
        control_value=0,
        treatment_value=hfv_value,
        method_params={"init_params":{'model_y': RandomForestRegressor(**param_Y),
                                      "model_t": RandomForestRegressor(**param_T),  
                                      "model_final": LassoCV(),  
                                      "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                     }}
    )
    causal_effect = estimate.value
    
    predicted_biogas_production_hfv = baseline_biogas_production + causal_effect

    results.append({
        "HFV": hfv_value,
        "Predicted_Biogas_Production": predicted_biogas_production_hfv 
    })

results_df = pd.DataFrame(results)
print(results_df)
results_df.to_excel("HFV_Predictions.xlsx", index=False)

print("反事实预测结果已保存为 Excel 文件。")

######----------HRT-----------######

import dowhy
from dowhy import CausalModel
import pandas as pd

file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['Biogas_Production']]  
T = dataset[['HRT']]  
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'Addition', 'Temperature', 'OLR', 'HFV', 'pH', 'TAN']]  
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]  

causal_graph = """
digraph {
    HRT -> Biogas_Production;
    PWFV -> HRT;
    FW -> HRT;
    ICO -> HRT;
    AO -> HRT;
    Manure -> HRT;
    Percolate -> HRT;
    Temperature -> HRT;
    OLR -> HRT;
    HFV -> HRT;
    pH -> HRT;
    TAN -> HRT;
    TSf -> HRT;
    VSf -> HRT;
    TSo -> HRT;
    VSo -> HRT;
    TSf -> Biogas_Production;
    VSf -> Biogas_Production;
    TSo -> Biogas_Production;
    VSo -> Biogas_Production;
}
"""
# I. Create a causal model from the data and domain knowledge.
model = CausalModel(
    data=dataset,
    treatment='HRT',
    outcome='Biogas_Production',
    common_causes=['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'Addition', 'Temperature', 'OLR', 'HFV', 'pH', 'TAN', 'TSf', 'VSf', 'TSo', 'VSo'],
    graph=causal_graph
)

model.view_model()
# layout="dot"
from IPython.display import Image, display
display(Image(filename="causal_model.png"))

# II. Identify causal effect and return target estimands
identified_estimand = model.identify_effect()
print(identified_estimand)

# III. Estimate the target estimand using a statistical method.
'''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
if (X == 0).all().all():
    Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
    X_train_combined =W_train
else:
    print(1)
    Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
    X_train_combined = np.hstack((X_train, W_train))
'''Hyperparameter optimization'''
param_Y = {}
param_T = {}
'''Hyperparameter optimization'''
param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "RandomForest")

### Calculate the baseline biogas production (HRT = 0)   ###
baseline_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=0,  
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  "model_t": RandomForestRegressor(**param_T),
                                  "model_final": LassoCV(),
                                  "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                 }}
)

baseline_biogas_production = baseline_estimate.value 
print("Baseline Biogas Production (HRT = 0):", baseline_biogas_production)

dml_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=1,
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  'model_t': RandomForestRegressor(**param_T),
                                  "model_final":LassoCV(),
                                  'featurizer':PolynomialFeatures(degree=2, include_bias=True)
                                               }}
)

print("DML ATE Estimate:", dml_estimate)

# IV. Refute the obtained estimate using multiple robustness checks.
refute_results = model.refute_estimate(identified_estimand, dml_estimate,
                                       method_name="placebo_treatment_refuter")
print(refute_results)

# Iterate over the range of HRT from 0 to 900 and perform intervention analysis
results = []
for hrt_value in range(0, 901, 1):  
    estimate = model.estimate_effect(
        identified_estimand,
        method_name="backdoor.econml.dml.DML",
        control_value=0,
        treatment_value=hrt_value,
        method_params={"init_params":{'model_y': RandomForestRegressor(**param_Y),
                                      "model_t": RandomForestRegressor(**param_T),  
                                      "model_final": LassoCV(),  
                                      "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                     }}
    )
    causal_effect = estimate.value
    
    predicted_biogas_production_hrt = baseline_biogas_production + causal_effect

    results.append({
        "HRT": hrt_value,
        "Predicted_Biogas_Production": predicted_biogas_production_hrt 
    })

results_df = pd.DataFrame(results)
print(results_df)
results_df.to_excel("HRT_Predictions.xlsx", index=False)

print("反事实预测结果已保存为 Excel 文件。")

######----------OLR-----------######

import dowhy
from dowhy import CausalModel
import pandas as pd

file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['Biogas_Production']]  
T = dataset[['OLR']]  
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'Addition', 'HRT', 'Temperature', 'HFV', 'pH', 'TAN']]  
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]  

causal_graph = """
digraph {
    OLR -> Biogas_Production;
    PWFV -> OLR;
    FW -> OLR;
    ICO -> OLR;
    AO -> OLR;
    Manure -> OLR;
    Percolate -> OLR;
    HRT -> OLR;
    Temperature -> OLR;
    HFV -> OLR;
    pH -> OLR;
    TAN -> OLR;
    TSf -> OLR;
    VSf -> OLR;
    TSo -> OLR;
    VSo -> OLR;
    TSf -> Biogas_Production;
    VSf -> Biogas_Production;
    TSo -> Biogas_Production;
    VSo -> Biogas_Production;
}
"""
# I. Create a causal model from the data and domain knowledge.
model = CausalModel(
    data=dataset,
    treatment='OLR',
    outcome='Biogas_Production',
    common_causes=['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'Addition', 'HRT', 'Temperature', 'HFV', 'pH', 'TAN', 'TSf', 'VSf', 'TSo', 'VSo'],
    graph=causal_graph
)

model.view_model()
# layout="dot"
from IPython.display import Image, display
display(Image(filename="causal_model.png"))

# II. Identify causal effect and return target estimands
identified_estimand = model.identify_effect()
print(identified_estimand)

# III. Estimate the target estimand using a statistical method.
'''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
if (X == 0).all().all():
    Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
    X_train_combined =W_train
else:
    print(1)
    Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
    X_train_combined = np.hstack((X_train, W_train))
'''Hyperparameter optimization'''
param_Y = {}
param_T = {}
'''Hyperparameter optimization'''
param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "RandomForest")

### Calculate the baseline biogas production (OLR = 0)   ###
baseline_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=0,  
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  "model_t": RandomForestRegressor(**param_T),
                                  "model_final": LassoCV(),
                                  "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                 }}
)

baseline_biogas_production = baseline_estimate.value  
print("Baseline Biogas Production (OLR = 0):", baseline_biogas_production)

dml_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=1,
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  'model_t': RandomForestRegressor(**param_T),
                                  "model_final":LassoCV(),
                                  'featurizer':PolynomialFeatures(degree=2, include_bias=True)
                                               }}
)

print("DML ATE Estimate:", dml_estimate)

# IV. Refute the obtained estimate using multiple robustness checks.
refute_results = model.refute_estimate(identified_estimand, dml_estimate,
                                       method_name="placebo_treatment_refuter")
print(refute_results)

# Iterate over the range of OLR from 0 to 3.5 and perform intervention analysis
results = []
for olr_value in np.arange(0, 3.6, 0.01):  
    estimate = model.estimate_effect(
        identified_estimand,
        method_name="backdoor.econml.dml.DML",
        control_value=0,
        treatment_value=[olr_value],
        method_params={"init_params":{'model_y': RandomForestRegressor(**param_Y),
                                      "model_t": RandomForestRegressor(**param_T),  
                                      "model_final": LassoCV(),  
                                      "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                     }}
    )
    causal_effect = estimate.value

    predicted_biogas_production_olr = baseline_biogas_production + causal_effect

    results.append({
        "OLR": olr_value,
        "Predicted_Biogas_Production": predicted_biogas_production_olr 
    })

results_df = pd.DataFrame(results)
print(results_df)
results_df.to_excel("OLR_Predictions.xlsx", index=False)

print("反事实预测结果已保存为 Excel 文件。")

######----------pH-----------######

import dowhy
from dowhy import CausalModel
import pandas as pd

file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['Biogas_Production']]  
T = dataset[['pH']]  
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'Addition', 'HRT', 'Temperature', 'HFV', 'OLR', 'TAN']]  
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]  

causal_graph = """
digraph {
    pH -> Biogas_Production;
    PWFV -> pH;
    FW -> pH;
    ICO -> pH;
    AO -> pH;
    Manure -> pH;
    Percolate -> pH;
    HRT -> pH;
    Temperature -> pH;
    HFV -> pH;
    OLR -> pH;
    TAN -> pH;
    TSf -> pH;
    VSf -> pH;
    TSo -> pH;
    VSo -> pH;
    TSf -> Biogas_Production;
    VSf -> Biogas_Production;
    TSo -> Biogas_Production;
    VSo -> Biogas_Production;
}
"""
# I. Create a causal model from the data and domain knowledge.
model = CausalModel(
    data=dataset,
    treatment='pH',
    outcome='Biogas_Production',
    common_causes=['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'Addition', 'HRT', 'Temperature', 'HFV', 'OLR', 'TAN', 'TSf', 'VSf', 'TSo', 'VSo'],
    graph=causal_graph
)

model.view_model()
# layout="dot"
from IPython.display import Image, display
display(Image(filename="causal_model.png"))

# II. Identify causal effect and return target estimands
identified_estimand = model.identify_effect()
print(identified_estimand)

# III. Estimate the target estimand using a statistical method.
'''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
if (X == 0).all().all():
    Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
    X_train_combined =W_train
else:
    print(1)
    Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
    X_train_combined = np.hstack((X_train, W_train))
'''Hyperparameter optimization'''
param_Y = {}
param_T = {}
'''Hyperparameter optimization'''
param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "RandomForest")


### Calculate the baseline biogas production (pH = 0)   ###
baseline_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=0, 
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  "model_t": RandomForestRegressor(**param_T),
                                  "model_final": LassoCV(),
                                  "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                 }}
)

baseline_biogas_production = baseline_estimate.value  
print("Baseline Biogas Production (pH = 0):", baseline_biogas_production)

dml_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=1,
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  'model_t': RandomForestRegressor(**param_T),
                                  "model_final":LassoCV(),
                                  'featurizer':PolynomialFeatures(degree=2, include_bias=True)
                                               }}
)

print("DML ATE Estimate:", dml_estimate)

# IV. Refute the obtained estimate using multiple robustness checks.
refute_results = model.refute_estimate(identified_estimand, dml_estimate,
                                       method_name="placebo_treatment_refuter")
print(refute_results)

# Iterate over the range of pH from 7.2 to 7.8 and perform intervention analysis
results = []
for ph_value in np.arange(7.2, 7.9, 0.01):  
    estimate = model.estimate_effect(
        identified_estimand,
        method_name="backdoor.econml.dml.DML",
        control_value=0,
        treatment_value=[ph_value],
        method_params={"init_params":{'model_y': RandomForestRegressor(**param_Y),
                                      "model_t": RandomForestRegressor(**param_T),  
                                      "model_final": LassoCV(),  
                                      "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                     }}
    )
    causal_effect = estimate.value
    
    predicted_biogas_production_ph = baseline_biogas_production + causal_effect

    results.append({
        "pH": ph_value,
        "Predicted_Biogas_Production": predicted_biogas_production_ph 
    })

results_df = pd.DataFrame(results)
print(results_df)
results_df.to_excel("pH_Predictions.xlsx", index=False)

print("反事实预测结果已保存为 Excel 文件。")

######----------Temperature_high-----------######

import dowhy
from dowhy import CausalModel
import pandas as pd

file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['Biogas_Production']]  
T = dataset[['Temperature']]  
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'Addition', 'HRT', 'OLR', 'HFV', 'pH', 'TAN']]  
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]  

causal_graph = """
digraph {
    Temperature -> Biogas_Production;
    PWFV -> Temperature;
    FW -> Temperature;
    ICO -> Temperature;
    AO -> Temperature;
    Manure -> Temperature;
    Percolate -> Temperature;
    HRT -> Temperature;
    OLR -> Temperature;
    HFV -> Temperature;
    pH -> Temperature;
    TAN -> Temperature;
    TSf -> Temperature;
    VSf -> Temperature;
    TSo -> Temperature;
    VSo -> Temperature;
    TSf -> Biogas_Production;
    VSf -> Biogas_Production;
    TSo -> Biogas_Production;
    VSo -> Biogas_Production;
}
"""
# I. Create a causal model from the data and domain knowledge.
model = CausalModel(
    data=dataset,
    treatment='Temperature',
    outcome='Biogas_Production',
    common_causes=['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'Addition', 'HRT', 'OLR', 'HFV', 'pH', 'TAN', 'TSf', 'VSf', 'TSo', 'VSo'],
    graph=causal_graph
)

model.view_model()
# layout="dot"
from IPython.display import Image, display
display(Image(filename="causal_model.png"))

# II. Identify causal effect and return target estimands
identified_estimand = model.identify_effect()
print(identified_estimand)

# III. Estimate the target estimand using a statistical method.
'''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
if (X == 0).all().all():
    Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
    X_train_combined =W_train
else:
    print(1)
    Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
    X_train_combined = np.hstack((X_train, W_train))
'''Hyperparameter optimization'''
param_Y = {}
param_T = {}
'''Hyperparameter optimization'''
param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "RandomForest")

### Calculate the baseline biogas production (Temperature_high = 0)   ###########################
baseline_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=0,  
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  "model_t": RandomForestRegressor(**param_T),
                                  "model_final": LassoCV(),
                                  "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                 }}
)

baseline_biogas_production = baseline_estimate.value  
print("Baseline Biogas Production (Temperature_high = 0):", baseline_biogas_production)

dml_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=1,
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  'model_t': RandomForestRegressor(**param_T),
                                  "model_final":LassoCV(),
                                  'featurizer':PolynomialFeatures(degree=2, include_bias=True)
                                               }}
)

print("DML ATE Estimate:", dml_estimate)

# IV. Refute the obtained estimate using multiple robustness checks.
refute_results = model.refute_estimate(identified_estimand, dml_estimate,
                                       method_name="placebo_treatment_refuter")
print(refute_results)

# Iterate over the range of Temperature from 50 to 56 and perform intervention analysis
results = []
for Temperature_high_value in np.arange(50, 56, 0.1):  
    estimate = model.estimate_effect(
        identified_estimand,
        method_name="backdoor.econml.dml.DML",
        control_value=0,
        treatment_value=[Temperature_high_value],
        method_params={"init_params":{'model_y': RandomForestRegressor(**param_Y),
                                      "model_t": RandomForestRegressor(**param_T),  
                                      "model_final": LassoCV(),  
                                      "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                     }}
    )
    causal_effect = estimate.value
    
    predicted_biogas_production_Temperature_high = baseline_biogas_production + causal_effect

    results.append({
        "HFV": Temperature_high_value,
        "Predicted_Biogas_Production": predicted_biogas_production_Temperature_high 
    })

results_df = pd.DataFrame(results)
print(results_df)
results_df.to_excel("Temperature_high_Predictions.xlsx", index=False)

print("反事实预测结果已保存为 Excel 文件。")

######----------Temperature_low-----------######

import dowhy
from dowhy import CausalModel
import pandas as pd

file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['Biogas_Production']]  
T = dataset[['Temperature']] 
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'Addition', 'HRT', 'OLR', 'HFV', 'pH', 'TAN']]  
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]  

causal_graph = """
digraph {
    Temperature -> Biogas_Production;
    PWFV -> Temperature;
    FW -> Temperature;
    ICO -> Temperature;
    AO -> Temperature;
    Manure -> Temperature;
    Percolate -> Temperature;
    HRT -> Temperature;
    OLR -> Temperature;
    HFV -> Temperature;
    pH -> Temperature;
    TAN -> Temperature;
    TSf -> Temperature;
    VSf -> Temperature;
    TSo -> Temperature;
    VSo -> Temperature;
    TSf -> Biogas_Production;
    VSf -> Biogas_Production;
    TSo -> Biogas_Production;
    VSo -> Biogas_Production;
}
"""
# I. Create a causal model from the data and domain knowledge.
model = CausalModel(
    data=dataset,
    treatment='Temperature',
    outcome='Biogas_Production',
    common_causes=['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'TFV', 'Addition', 'HRT', 'OLR', 'HFV', 'pH', 'TAN', 'TSf', 'VSf', 'TSo', 'VSo'],
    graph=causal_graph
)

model.view_model()
# layout="dot"
from IPython.display import Image, display
display(Image(filename="causal_model.png"))

# II. Identify causal effect and return target estimands
identified_estimand = model.identify_effect()
print(identified_estimand)

# III. Estimate the target estimand using a statistical method.
'''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
if (X == 0).all().all():
    Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
    X_train_combined =W_train
else:
    print(1)
    Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
    X_train_combined = np.hstack((X_train, W_train))
'''Hyperparameter optimization'''
param_Y = {}
param_T = {}
'''Hyperparameter optimization'''
param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "RandomForest")

### Calculate the baseline biogas production (Temperature_low = 0)   ###
baseline_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=0,  
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  "model_t": RandomForestRegressor(**param_T),
                                  "model_final": LassoCV(),
                                  "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                 }}
)

baseline_biogas_production = baseline_estimate.value 
print("Baseline Biogas Production (Temperature_low = 0):", baseline_biogas_production)

dml_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=1,
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  'model_t': RandomForestRegressor(**param_T),
                                  "model_final":LassoCV(),
                                  'featurizer':PolynomialFeatures(degree=2, include_bias=True)
                                               }}
)
print("DML ATE Estimate:", dml_estimate)

# IV. Refute the obtained estimate using multiple robustness checks.
refute_results = model.refute_estimate(identified_estimand, dml_estimate,
                                       method_name="placebo_treatment_refuter")
print(refute_results)

# Iterate over the range of Temperature from 35 to 38 and perform intervention analysis
results = []
for Temperature_low_value in np.arange(35, 39, 0.1):  
    estimate = model.estimate_effect(
        identified_estimand,
        method_name="backdoor.econml.dml.DML",
        control_value=0,
        treatment_value=[Temperature_low_value],
        method_params={"init_params":{'model_y': RandomForestRegressor(**param_Y),
                                      "model_t": RandomForestRegressor(**param_T),  
                                      "model_final": LassoCV(),  
                                      "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                     }}
    )
    causal_effect = estimate.value

    predicted_biogas_production_Temperature_low = baseline_biogas_production + causal_effect

    results.append({
        "HFV": Temperature_low_value,
        "Predicted_Biogas_Production": predicted_biogas_production_Temperature_low 
    })

results_df = pd.DataFrame(results)
print(results_df)
results_df.to_excel("Temperature_low_Predictions.xlsx", index=False)

print("反事实预测结果已保存为 Excel 文件。")

######----------TFV-----------######

import dowhy
from dowhy import CausalModel
import pandas as pd

file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['Biogas_Production']]  
T = dataset[['TFV']]  
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'HFV', 'Addition', 'Temperature', 'OLR', 'HRT', 'pH', 'TAN']]  
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]  

causal_graph = """
digraph {
    TFV -> Biogas_Production;
    PWFV -> TFV;
    FW -> TFV;
    ICO -> TFV;
    AO -> TFV;
    Manure -> TFV;
    Percolate -> TFV;
    Temperature -> TFV;
    OLR -> TFV;
    HRT -> TFV;
    pH -> TFV;
    TAN -> TFV;
    TSf -> TFV;
    VSf -> TFV;
    TSo -> TFV;
    VSo -> TFV;
    TSf -> Biogas_Production;
    VSf -> Biogas_Production;
    TSo -> Biogas_Production;
    VSo -> Biogas_Production;
}
"""
# I. Create a causal model from the data and domain knowledge.
model = CausalModel(
    data=dataset,
    treatment='TFV',
    outcome='Biogas_Production',
    common_causes=['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'VFA', 'HFV', 'Addition', 'Temperature', 'OLR', 'HRT', 'pH', 'TAN', 'TSf', 'VSf', 'TSo', 'VSo'],
    graph=causal_graph
)

model.view_model()
# layout="dot"
from IPython.display import Image, display
display(Image(filename="causal_model.png"))

# II. Identify causal effect and return target estimands
identified_estimand = model.identify_effect()
print(identified_estimand)

# III. Estimate the target estimand using a statistical method.
'''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
if (X == 0).all().all():
    Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
    X_train_combined =W_train
else:
    print(1)
    Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
    X_train_combined = np.hstack((X_train, W_train))
'''Hyperparameter optimization'''
param_Y = {}
param_T = {}
'''Hyperparameter optimization'''
param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "RandomForest")

### Calculate the baseline biogas production (TFV = 0)   ###########################
baseline_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=0,  
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  "model_t": RandomForestRegressor(**param_T),
                                  "model_final": LassoCV(),
                                  "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                 }}
)

baseline_biogas_production = baseline_estimate.value  
print("Baseline Biogas Production (TFV = 0):", baseline_biogas_production)

dml_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=1,
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  'model_t': RandomForestRegressor(**param_T),
                                  "model_final":LassoCV(),
                                  'featurizer':PolynomialFeatures(degree=2, include_bias=True)
                                               }}
)

print("DML ATE Estimate:", dml_estimate)

# IV. Refute the obtained estimate using multiple robustness checks.
refute_results = model.refute_estimate(identified_estimand, dml_estimate,
                                       method_name="placebo_treatment_refuter")
print(refute_results)

# Iterate over the range of TFV from 0 to 400 and perform intervention analysis
results = []
for tfv_value in np.arange(0, 401, 0.5):  
    estimate = model.estimate_effect(
        identified_estimand,
        method_name="backdoor.econml.dml.DML",
        control_value=0,
        treatment_value=[tfv_value],
        method_params={"init_params":{'model_y': RandomForestRegressor(**param_Y),
                                      "model_t": RandomForestRegressor(**param_T),  
                                      "model_final": LassoCV(),  
                                      "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                     }}
    )
    causal_effect = estimate.value
    
    predicted_biogas_production_tfv = baseline_biogas_production + causal_effect

    results.append({
        "TFV": tfv_value,
        "Predicted_Biogas_Production": predicted_biogas_production_tfv 
    })

results_df = pd.DataFrame(results)
print(results_df)
results_df.to_excel("TFV_Predictions.xlsx", index=False)

print("反事实预测结果已保存为 Excel 文件。")


######----------VFA-----------######

import dowhy
from dowhy import CausalModel
import pandas as pd

file_path = 'jz_data_knn.xlsx'
dataset = pd.read_excel(file_path, index_col=None, keep_default_na=True)

Y = dataset[['Biogas_Production']]  
T = dataset[['VFA']]  
X = dataset[['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'HFV', 'Addition', 'Temperature', 'OLR', 'HRT', 'pH', 'TAN']]  
W = dataset[['TSf', 'VSf', 'TSo', 'VSo']]  

causal_graph = """
digraph {
    VFA -> Biogas_Production;
    PWFV -> VFA;
    FW -> VFA;
    ICO -> VFA;
    AO -> VFA;
    Manure -> VFA;
    Percolate -> VFA;
    Temperature -> VFA;
    OLR -> VFA;
    HRT -> VFA;
    pH -> VFA;
    TAN -> VFA;
    TSf -> VFA;
    VSf -> VFA;
    TSo -> VFA;
    VSo -> VFA;
    TSf -> Biogas_Production;
    VSf -> Biogas_Production;
    TSo -> Biogas_Production;
    VSo -> Biogas_Production;
}
"""
# Create a causal model from the data and updated domain knowledge
model = CausalModel(
    data=dataset,
    treatment='VFA',
    outcome='Biogas_Production',
    common_causes=['PWFV', 'FW', 'ICO', 'AO', 'Manure', 'Percolate', 'TFV', 'HFV', 'Addition', 'Temperature', 'OLR', 'HRT', 'pH', 'TAN', 'TSf', 'VSf', 'TSo', 'VSo'],
    graph=causal_graph
)

model.view_model()
# layout="dot"
from IPython.display import Image, display
display(Image(filename="causal_model.png"))

# Identify causal effect and return target estimands
identified_estimand = model.identify_effect()
print(identified_estimand)

# III. Estimate the target estimand using a statistical method.
'''splitting: Type of splitting ('Random splitting' or 'Grouped random splitting')'''
if (X == 0).all().all():
    Y_train, Y_test,T_train, T_test,W_train, W_test = train_test_split(Y,T,W, test_size=0.2, random_state=0)
    X_train_combined =W_train
else:
    print(1)
    Y_train, Y_test,T_train, T_test, X_train, X_test, W_train, W_test = train_test_split(Y,T,X,W, test_size=0.2, random_state=0)
    X_train_combined = np.hstack((X_train, W_train))
'''Hyperparameter optimization'''
param_Y = {}
param_T = {}
'''Hyperparameter optimization'''
param_Y, best_score_Y = optimizer_optuna(X_train_combined, Y_train, 30, "RandomForest")
param_T, best_score_T = optimizer_optuna(X_train_combined, T_train, 30, "CatBoost")

### Calculate the baseline biogas production (VFA = 0)   ###
baseline_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=0,  
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  "model_t": CatBoostRegressor(**param_T),
                                  "model_final": LassoCV(),
                                  "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                 }}
)

baseline_biogas_production = baseline_estimate.value  
print("Baseline Biogas Production (VFA = 0):", baseline_biogas_production)

dml_estimate = model.estimate_effect(
    identified_estimand,
    method_name="backdoor.econml.dml.DML",
    control_value=0,
    treatment_value=1,
    method_params={"init_params":{'model_y':RandomForestRegressor(**param_Y),
                                  'model_t': CatBoostRegressor(**param_T),
                                  "model_final":LassoCV(),
                                  'featurizer':PolynomialFeatures(degree=2, include_bias=True)
                                               }}
)

print("DML ATE Estimate:", dml_estimate)

# Iterate over the range of VFA from 0 to 3000 and perform intervention analysis
results = []
for vfa_value in range(0, 3001, 1):  
    estimate = model.estimate_effect(
        identified_estimand,
        method_name="backdoor.econml.dml.DML",
        control_value=0,
        treatment_value=vfa_value,
        method_params={"init_params":{'model_y': RandomForestRegressor(**param_Y),
                                      "model_t": CatBoostRegressor(**param_T),  
                                      "model_final": LassoCV(),  
                                      "featurizer": PolynomialFeatures(degree=2, include_bias=True)
                                     }}
    )
    causal_effect = estimate.value
    predicted_biogas_production_vfa =  baseline_biogas_production + causal_effect  

    results.append({
        "VFA": vfa_value,
        "Predicted_Biogas_Production": predicted_biogas_production_vfa
    })

results_df = pd.DataFrame(results)
print(results_df)
results_df.to_excel("VFA_Predictions.xlsx", index=False)

print("反事实预测结果已保存为 Excel 文件。")

